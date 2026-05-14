"""Write results to multi-sheet xlsx — one sheet per pipeline module."""
from __future__ import annotations
from pathlib import Path
from datetime import datetime
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .config import PipelineConfig
from .utils import _scalar


_BANNER  = PatternFill("solid", fgColor="1F3864")
_SUB     = PatternFill("solid", fgColor="2E5395")
_HDR     = PatternFill("solid", fgColor="D9E2F3")
_SEC_IN  = PatternFill("solid", fgColor="F2F7FC")
_SEC_OUT = PatternFill("solid", fgColor="FFF7E1")
_ALT     = PatternFill("solid", fgColor="F8F8F8")
_BORDER  = Border(left=Side(style="thin", color="C0C0C0"),
                  right=Side(style="thin", color="C0C0C0"),
                  top=Side(style="thin", color="C0C0C0"),
                  bottom=Side(style="thin", color="C0C0C0"))
_FT_TITLE = Font(name="Arial", size=14, bold=True, color="FFFFFF")
_FT_SUB   = Font(name="Arial", size=11, bold=True, color="FFFFFF")
_FT_HDR   = Font(name="Arial", size=10, bold=True, color="1F3864")
_FT_BODY  = Font(name="Arial", size=10)


def _banner(ws, title, sub, ncols=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = _FT_TITLE; c.fill = _BANNER
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=sub)
    c.font = _FT_SUB; c.fill = _SUB
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18


def _section(ws, row, label, ncols=8, is_input=False):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(name="Arial", size=10, bold=True, color="1F3864")
    c.fill = _SEC_IN if is_input else _SEC_OUT
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    return row + 1


def _kv(ws, start_row, kv):
    r = start_row
    for k, v in kv.items():
        ws.cell(row=r, column=1, value=str(k)).font = _FT_HDR
        c = ws.cell(row=r, column=2, value=_scalar(v))
        c.font = _FT_BODY; c.alignment = Alignment(horizontal="left")
        r += 1
    return r


def _table(ws, start_row, df):
    if df is None or len(df) == 0:
        ws.cell(row=start_row, column=1, value="(no rows)").font = _FT_BODY
        return start_row + 1
    for j, col in enumerate(df.columns):
        c = ws.cell(row=start_row, column=1+j, value=str(col))
        c.font = _FT_HDR; c.fill = _HDR
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER
    for i, (_, row) in enumerate(df.iterrows()):
        for j, col in enumerate(df.columns):
            c = ws.cell(row=start_row+1+i, column=1+j, value=_scalar(row[col]))
            c.font = _FT_BODY; c.border = _BORDER
            if i % 2 == 1: c.fill = _ALT
    return start_row + 1 + len(df)


def _autosize(ws, min_w=10, max_w=40):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 8
        for cell in ws[letter]:
            try:
                v = str(cell.value) if cell.value is not None else ""
                if len(v) > max_len: max_len = len(v)
            except Exception: pass
        ws.column_dimensions[letter].width = max(min_w, min(max_w, max_len + 2))


def _sheet_run_summary(wb, results, cfg, source):
    ws = wb.create_sheet("00_Run_Summary")
    _banner(ws, "PV Diagnostics — Run Summary",
            f"Plant: {cfg.site.name}  |  Run: {datetime.now():%Y-%m-%d %H:%M}  |  Source: {Path(source).name}", 6)
    pl = results.get("plant_losses", {})
    pm = results.get("plant_meta", {})
    bi = results.get("baseline_info", {})
    cur = cfg.site.currency
    r = 4
    r = _section(ws, r, "▸ Plant context", 6)
    r = _kv(ws, r, {
        "Plant name": cfg.site.name,
        "Latitude":  f"{cfg.site.lat:.4f}",
        "Longitude": f"{cfg.site.lon:.4f}",
        "Timezone":  cfg.site.tz,
        "Commissioning date": str(cfg.plant.commissioning_date),
        "Default azimuth (°)": cfg.plant.default_azimuth,
        "Default tilt (°)": cfg.plant.default_tilt,
        f"Tariff ({cur}/kWh)": cfg.site.tariff,
        "# Inverters": len(pm.get("inverters", [])),
        "# Strings": pm.get("total_strings", 0),
        "Period start": str(pm.get("ts_min", "")),
        "Period end":   str(pm.get("ts_max", "")),
        "Sampling (min)": pm.get("freq_min", ""),
    })
    r += 1
    r = _section(ws, r, "▸ Verdict distribution", 6)
    r = _kv(ws, r, results.get("verdict_counts", {}))
    r += 1
    r = _section(ws, r, "▸ Plant-wide losses", 6)
    r = _kv(ws, r, {
        "Period (days)": pl.get("period_days", 0),
        f"Soiling kWh": f"{pl.get('soiling_kwh',0):,.1f}",
        f"Soiling {cur}": f"{pl.get('soiling_pkr',0):,.0f}",
        f"Curtailment kWh": f"{pl.get('curtailment_kwh',0):,.1f}",
        f"Curtailment {cur}": f"{pl.get('curtailment_pkr',0):,.0f}",
        f"Total avoidable kWh": f"{pl.get('total_avoidable_kwh',0):,.1f}",
        f"Total avoidable {cur}": f"{pl.get('total_avoidable_pkr',0):,.0f}",
        f"Annualised kWh (x365/period)": f"{pl.get('annualised_kwh',0):,.0f}",
        f"Annualised {cur}": f"{pl.get('annualised_pkr',0):,.0f}",
    })
    r += 1
    r = _section(ws, r, "▸ Degradation baseline applied", 6)
    r = _kv(ws, r, {
        "Module technology": bi.get("technology", ""),
        "Years since commissioning": f"{bi.get('years',0):.2f}",
        "Annual rate (frac/yr)": bi.get("annual_rate", ""),
        "LID applied (frac)": f"{bi.get('lid_applied',0):.4f}",
        "Degradation applied (frac)": f"{bi.get('deg_applied',0):.4f}",
        "Baseline used": f"{bi.get('baseline',1):.4f}",
        "Apply correction": cfg.apply_degradation_correction,
    })
    _autosize(ws, 22, 46)


def _sheet_metadata(wb, results, cfg):
    ws = wb.create_sheet("01_Plant_Metadata")
    _banner(ws, "Plant Metadata", "Ingestion outputs + PK-default substitutions", 6)
    pm = results.get("plant_meta", {})
    rsl = pm.get("plant_resolved", {})
    r = 4
    r = _section(ws, r, "▸ INPUTS (from xlsx + Metadata sheet)", 6, is_input=True)
    r = _kv(ws, r, {
        "plant_name": rsl.get("plant_name", ""),
        "lat": rsl.get("lat", ""),
        "lon": rsl.get("lon", ""),
        "tariff": rsl.get("tariff", ""),
        "commissioning_date": str(rsl.get("commissioning_date", "")),
        "p_ac_max_kw": rsl.get("p_ac_max_kw", ""),
        "technology": rsl.get("technology", ""),
    })
    r += 1
    r = _section(ws, r, "▸ Substitution notes (defaults applied)", 6)
    for n in pm.get("substitution_notes", []) or ["No substitutions needed"]:
        ws.cell(row=r, column=1, value=f"• {n}").font = _FT_BODY
        r += 1
    r += 1
    r = _section(ws, r, "▸ Discovered structure", 6)
    r = _kv(ws, r, {
        "Plants": ", ".join(pm.get("plants", [])),
        "Inverters": ", ".join(pm.get("inverters", [])),
        "Total strings": pm.get("total_strings", 0),
        "Period start": str(pm.get("ts_min", "")),
        "Period end":   str(pm.get("ts_max", "")),
        "Sampling freq (min)": pm.get("freq_min", ""),
        "n_intervals": pm.get("n_intervals", ""),
        "Azimuth rows defaulted": pm.get("azimuth_filled_rows", 0),
        "Tilt rows defaulted": pm.get("tilt_filled_rows", 0),
    })
    _autosize(ws, 22, 60)


def _sheet_sufficiency(wb, results, cfg):
    ws = wb.create_sheet("02_Data_Sufficiency")
    _banner(ws, "Data Sufficiency", "Per-string availability & sufficiency gate", 8)
    rows = []
    for label, r in results["per_string"].items():
        dq = r.get("data_quality", {}) or {}
        rows.append(dict(string_label=label, verdict=r.get("sufficiency",""),
            reason=r.get("sufficiency_reason",""),
            avail_pct=dq.get("avail_pct"), curt_pct=dq.get("curt_pct"),
            fault_pct=dq.get("fault_pct"), max_gap_days=dq.get("max_gap_days"),
            n_daylight=dq.get("n_daylight"), n_ok=dq.get("n_ok")))
    _table(ws, 4, pd.DataFrame(rows)); _autosize(ws)


def _sheet_curt_detection(wb, results, cfg):
    ws = wb.create_sheet("03_Curtailment_Detection")
    _banner(ws, "Curtailment Detection", "Per-string state + statistical plateaus", 8)
    rows = []
    for label, r in results["per_string"].items():
        cs = r.get("curtailment_summary", {}) or {}
        rows.append(dict(string_label=label,
            n_curt_state=cs.get("n_curt_state",0),
            n_curt_stat=cs.get("n_curt_stat",0),
            n_curt_total=cs.get("n_curt_total",0),
            curt_pct=cs.get("curt_pct",np.nan),
            curt_hours_state=cs.get("curt_hours_state",np.nan),
            curt_hours_stat=cs.get("curt_hours_stat",np.nan),
            top_state_codes=str(cs.get("top_state_codes",""))))
    _table(ws, 4, pd.DataFrame(rows)); _autosize(ws)


def _sheet_curt_loss(wb, results, cfg):
    ws = wb.create_sheet("04_Curtailment_Loss")
    _banner(ws, "Curtailment Loss (kWh & PKR)",
            f"Tariff = {cfg.site.tariff:.1f} {cfg.site.currency}/kWh", 10)
    cur = cfg.site.currency
    rows = []
    for label, r in results["per_string"].items():
        cl = r.get("curt_loss", {}) or {}
        rows.append({"string_label":label,
            "n_curt_intervals":cl.get("n_curt_intervals",0),
            "total_curt_kwh":cl.get("total_curt_kwh",0.0),
            f"total_curt_{cur}":cl.get("total_curt_pkr",0.0),
            "period_days":cl.get("period_days",0),
            "annualised_kwh":cl.get("annualised_kwh",0.0),
            f"annualised_{cur}":cl.get("annualised_pkr",0.0),
            "method":cl.get("method",""),
            "explainability":cl.get("explainability","")})
    _table(ws, 4, pd.DataFrame(rows)); _autosize(ws, max_w=60)


def _sheet_plate(wb, results, cfg):
    ws = wb.create_sheet("05_Plate_Inference")
    _banner(ws, "Plate Inference", "STC nameplate inferred from clean data", 4)
    pi = results.get("plate_inferred", {}) or {}
    plate = results.get("plate", cfg.module)
    r = 4
    r = _section(ws, r, "▸ INPUTS", 4, is_input=True)
    r = _kv(ws, r, {
        "Technology": getattr(plate, "technology", "mono-c-Si"),
        "Cells in series": getattr(plate, "cells_in_series", 144),
        "Modules per string": getattr(plate, "n_modules", 22),
    })
    r += 1
    r = _section(ws, r, "▸ OUTPUTS (inferred)", 4)
    r = _kv(ws, r, {
        "Voc_stc (V)": plate.voc_stc, "Vmp_stc (V)": plate.vmp_stc,
        "Isc_stc (A)": plate.isc_stc, "Imp_stc (A)": plate.imp_stc,
        "alpha_isc": plate.alpha_isc, "beta_voc": plate.beta_voc,
        "gamma_pmp": plate.gamma_pmp,
        "Pmp string STC (W)": plate.pmp_str_stc,
        "Notes": pi.get("notes", ""),
    })
    _autosize(ws, 24, 40)


def _sheet_sdm(wb, results, cfg):
    ws = wb.create_sheet("06_SDM_Fits")
    _banner(ws, "Single-Diode Model Fits", "Rs / Rsh / I0 / n + STC fingerprint", 10)
    rows = []
    for label, r in results["per_string"].items():
        s = r.get("sdm", {}) or {}
        sm = r.get("sdm_metrics", {}) or {}
        rows.append(dict(string_label=label,
            success=s.get("success",False), reason=s.get("reason",""),
            n_samples=s.get("n_samples",0),
            Rs_ohm=s.get("Rs"), Rsh_ohm=s.get("Rsh"),
            I0_A=s.get("I0"), n_diode=s.get("n"),
            voc_stc=sm.get("voc_stc"), isc_stc=sm.get("isc_stc"),
            ff=sm.get("ff"),
            voc_stc_ratio=sm.get("voc_stc_ratio"),
            isc_stc_ratio=sm.get("isc_stc_ratio"),
            ff_stc_ratio=sm.get("ff_stc_ratio")))
    _table(ws, 4, pd.DataFrame(rows)); _autosize(ws, max_w=60)


def _sheet_daily(wb, results, cfg):
    ws = wb.create_sheet("07_Daily_Metrics")
    _banner(ws, "Daily Metrics", "PR, NCI, NCI_corrected, asymmetry per day", 12)
    frames = []
    for label, r in results["per_string"].items():
        d = r.get("daily_df")
        if d is None or d.empty: continue
        d = d.copy(); d.insert(0, "string_label", label)
        frames.append(d)
    if frames:
        df = pd.concat(frames, ignore_index=True)
        _table(ws, 4, df)
    else:
        ws.cell(row=4, column=1, value="(no daily data)").font = _FT_BODY
    _autosize(ws, max_w=24)


def _sheet_degradation(wb, results, cfg):
    ws = wb.create_sheet("08_Degradation_Baseline")
    _banner(ws, "Degradation Baseline", "Age-corrected NCI baseline", 6)
    bi = results.get("baseline_info", {}) or {}
    r = 4
    r = _section(ws, r, "▸ INPUTS", 6, is_input=True)
    r = _kv(ws, r, {
        "Commissioning date": str(bi.get("commissioning_date","")),
        "Reference date": str(bi.get("reference_date","")),
        "Module technology": bi.get("technology",""),
        "Annual rate (frac/yr)": bi.get("annual_rate",""),
        "LID rate (first year)": bi.get("lid_rate",""),
        "Floor": bi.get("floor",""),
        "Apply correction": cfg.apply_degradation_correction,
    })
    r += 1
    r = _section(ws, r, "▸ OUTPUTS", 6)
    r = _kv(ws, r, {
        "Years since commissioning": f"{bi.get('years',0):.3f}",
        "LID applied (frac)": bi.get("lid_applied",0),
        "Degradation applied (frac)": bi.get("deg_applied",0),
        "Baseline used (NCI ref)": bi.get("baseline",1),
        "Explanation": bi.get("note",""),
    })
    _autosize(ws, 26, 50)


def _sheet_wash(wb, results, cfg):
    ws = wb.create_sheet("09_Wash_Events")
    _banner(ws, "Wash / Rain Recovery Events",
            "Step-ups in NCI after a downward trend", 12)
    frames = []
    for label, r in results["per_string"].items():
        w = r.get("wash", {}) or {}
        ev = w.get("events_df", pd.DataFrame())
        if ev is None or ev.empty: continue
        ev = ev.copy(); ev.insert(0, "string_label", label)
        frames.append(ev)
    if frames:
        _table(ws, 4, pd.concat(frames, ignore_index=True))
    else:
        ws.cell(row=4, column=1, value="(no wash events detected)").font = _FT_BODY
    _autosize(ws, max_w=30)


def _sheet_soiling(wb, results, cfg):
    ws = wb.create_sheet("10_Soiling_Trends")
    _banner(ws, "Soiling Trends", "Full-window vs Current-segment slopes", 10)
    rows = []
    for label, r in results["per_string"].items():
        sf = r.get("soiling_full", {}) or {}
        sc = r.get("soiling_current", {}) or {}
        rows.append(dict(string_label=label,
            srr_full_pct_per_day=sf.get("srr_pct_per_day"),
            ci_full_pct_per_day=sf.get("ci_pct_per_day"),
            soiling_loss_full_pct=sf.get("weighted_soiling_loss_pct"),
            median_recovery_depth_pct=sf.get("median_recovery_depth_pct"),
            n_segments=sf.get("n_segments",0),
            srr_current_pct_per_day=sc.get("srr_pct_per_day"),
            soiling_loss_current_pct=sc.get("weighted_soiling_loss_pct"),
            method=sf.get("method",""),
            explainability=(sf.get("explainability","") or "")[:300]))
    _table(ws, 4, pd.DataFrame(rows)); _autosize(ws, max_w=60)


def _sheet_clusters(wb, results, cfg):
    ws = wb.create_sheet("11_Orientation_Clusters")
    _banner(ws, "Orientation Clusters", "MPPT + (az, tilt) combined", 10)
    _table(ws, 4, results.get("cluster_table", pd.DataFrame()))
    _autosize(ws, max_w=30)


def _sheet_class(wb, results, cfg):
    ws = wb.create_sheet("12_Classification")
    _banner(ws, "Classification", "Per-string verdict + axes", 14)
    rows = []
    for label, r in results["per_string"].items():
        c = r.get("classification", {}) or {}
        a = c.get("axes", {}) or {}
        rows.append(dict(string_label=label,
            verdict=c.get("verdict",""),
            primary_axis=c.get("primary_axis",""),
            confidence=c.get("confidence",""),
            soiling_band=a.get("soiling_band",""),
            mean_nci_current=a.get("mean_nci_current"),
            wash_event_recovery=a.get("wash_event_recovery") or "",
            wash_event_cause=a.get("wash_event_cause") or "",
            obs_asymmetry=a.get("obs_asymmetry"),
            expected_asymmetry=a.get("expected_asymmetry"),
            excess_asymmetry=a.get("excess_asymmetry"),
            has_shading_flag=a.get("has_shading_flag",False),
            has_degradation_flag=a.get("has_degradation_flag",False),
            srr_current_pct_per_day=a.get("srr_current_pct_per_day"),
            n_days_current_segment=a.get("n_days_current_segment",0),
            explainability=(c.get("explainability","") or "")[:400]))
    _table(ws, 4, pd.DataFrame(rows)); _autosize(ws, max_w=60)


def _sheet_losses(wb, results, cfg):
    ws = wb.create_sheet("13_Losses")
    _banner(ws, "Losses — Soiling + Curtailment",
            f"Tariff = {cfg.site.tariff:.1f} {cfg.site.currency}/kWh", 10)
    cur = cfg.site.currency
    rows = []
    for label, r in results["per_string"].items():
        l = r.get("losses", {}) or {}
        rows.append({"string_label":label,
            "soiling_kwh":l.get("soiling_kwh",0),
            f"soiling_{cur}":l.get("soiling_pkr",0),
            "curtailment_kwh":l.get("curtailment_kwh",0),
            f"curtailment_{cur}":l.get("curtailment_pkr",0),
            "total_avoidable_kwh":l.get("total_avoidable_kwh",0),
            f"total_avoidable_{cur}":l.get("total_avoidable_pkr",0),
            "annualised_kwh":l.get("annualised_kwh",0),
            f"annualised_{cur}":l.get("annualised_pkr",0),
            "period_days":l.get("period_days",0)})
    df = pd.DataFrame(rows)
    if not df.empty:
        tot_row = {c: (df[c].sum() if pd.api.types.is_numeric_dtype(df[c]) else "")
                   for c in df.columns}
        tot_row["string_label"] = "PLANT TOTAL"
        df = pd.concat([df, pd.DataFrame([tot_row])], ignore_index=True)
    _table(ws, 4, df); _autosize(ws, max_w=30)


def _sheet_transients(wb, results, cfg):
    ws = wb.create_sheet("14_Transient_Events")
    _banner(ws, "Transient Daily Events", "Single-day anomalous dips", 10)
    frames = []
    for label, r in results["per_string"].items():
        t = r.get("transients")
        if t is None or t.empty: continue
        t = t.copy(); t.insert(0, "string_label", label)
        frames.append(t)
    if frames:
        _table(ws, 4, pd.concat(frames, ignore_index=True))
    else:
        ws.cell(row=4, column=1, value="(no transients detected)").font = _FT_BODY
    _autosize(ws, max_w=40)


def _sheet_per_string(wb, results, cfg):
    ws = wb.create_sheet("15_Per_String_Detail")
    _banner(ws, "Per-String Compact Detail", "One row per string", 20)
    cur = cfg.site.currency
    rows = []
    for label, r in results["per_string"].items():
        m = r.get("meta", {}) or {}
        c = r.get("classification", {}) or {}
        a = c.get("axes", {}) or {}
        sf = r.get("soiling_full", {}) or {}
        sc = r.get("soiling_current", {}) or {}
        l = r.get("losses", {}) or {}
        w = r.get("wash", {}) or {}
        cl = r.get("cluster", {}) or {}
        rows.append({"string_label":label,
            "inverter":m.get("inverter_id",""),
            "mppt":m.get("mppt_id",""),
            "azimuth":m.get("azimuth",np.nan),
            "tilt":m.get("tilt",np.nan),
            "full_cluster":cl.get("full_cluster",""),
            "sufficiency":r.get("sufficiency",""),
            "verdict":c.get("verdict",""),
            "confidence":c.get("confidence",""),
            "soiling_band":a.get("soiling_band",""),
            "mean_nci_current":a.get("mean_nci_current"),
            "srr_full_pct_per_day":sf.get("srr_pct_per_day"),
            "srr_current_pct_per_day":sc.get("srr_pct_per_day"),
            "n_wash_events":w.get("n_events",0),
            "most_recent_event_class":(w.get("most_recent_event") or {}).get("recovery_class",""),
            "soiling_kwh":l.get("soiling_kwh",0),
            f"soiling_{cur}":l.get("soiling_pkr",0),
            "curt_kwh":l.get("curtailment_kwh",0),
            f"curt_{cur}":l.get("curtailment_pkr",0),
            "total_avoidable_kwh":l.get("total_avoidable_kwh",0),
            f"total_{cur}":l.get("total_avoidable_pkr",0)})
    _table(ws, 4, pd.DataFrame(rows)); _autosize(ws, max_w=30)


def export_results_to_excel(results, out_path, source_file="", verbose=True):
    out_path = Path(out_path); out_path.parent.mkdir(parents=True, exist_ok=True)
    cfg = results["cfg"]
    wb = Workbook(); wb.remove(wb.active)
    _sheet_run_summary(wb, results, cfg, source_file)
    _sheet_metadata(wb, results, cfg)
    _sheet_sufficiency(wb, results, cfg)
    _sheet_curt_detection(wb, results, cfg)
    _sheet_curt_loss(wb, results, cfg)
    _sheet_plate(wb, results, cfg)
    _sheet_sdm(wb, results, cfg)
    _sheet_daily(wb, results, cfg)
    _sheet_degradation(wb, results, cfg)
    _sheet_wash(wb, results, cfg)
    _sheet_soiling(wb, results, cfg)
    _sheet_clusters(wb, results, cfg)
    _sheet_adaptive_baseline(wb, results, cfg)
    _sheet_class(wb, results, cfg)
    _sheet_losses(wb, results, cfg)
    _sheet_transients(wb, results, cfg)
    _sheet_per_string(wb, results, cfg)
    wb.save(out_path)
    if verbose:
        print(f"  Wrote {out_path}  ({len(wb.sheetnames)} sheets)")
    return str(out_path)


def _sheet_adaptive_baseline(wb, results, cfg):
    """Sheet 11B — Adaptive Baseline provenance (one row per string).

    Mandatory for industrial trust: auditors can trace every resolved clean
    reference back to its layer, source, and the reasons earlier layers were
    rejected.
    """
    ws = wb.create_sheet("11B_Adaptive_Baseline")
    _banner(ws,
            "Adaptive Baseline — Clean Reference Provenance",
            "Layer 1=per-string P95 | Layer 2=cluster | Layer 3=plate fallback",
            10)

    rows = []
    adaptive_results = results.get("adaptive_results", {})

    for label, r in results["per_string"].items():
        # Prefer the structured AdaptiveBaselineResult if available
        ar = r.get("adaptive_baseline") or adaptive_results.get(label)

        if ar is None:
            # adaptive was disabled or string had an error
            rows.append(dict(
                string_label=label,
                cluster_id="",
                layer=None,
                source="adaptive_disabled",
                value=None,
                explainability="adaptive_baseline_enabled=False",
                p95=None,
                p50=None,
                n_used=None,
                n_rain_events_in_window=None,
                baseline_disagreement_flag=None,
                baseline_disagreement_pp=None,
            ))
        else:
            axes = (r.get("classification") or {}).get("axes") or {}
            rows.append(dict(
                string_label=label,
                cluster_id=getattr(ar, "cluster_id", ""),
                layer=getattr(ar, "layer", None),
                source=getattr(ar, "source", ""),
                value=getattr(ar, "value", None),
                explainability=getattr(ar, "explainability", ""),
                p95=getattr(ar, "p95", None),
                p50=getattr(ar, "p50", None),
                n_used=getattr(ar, "n_used", None),
                n_rain_events_in_window=getattr(ar, "n_rain_events_in_window", None),
                baseline_disagreement_flag=axes.get("baseline_disagreement_flag"),
                baseline_disagreement_pp=axes.get("baseline_disagreement_pp"),
            ))

    _table(ws, 4, pd.DataFrame(rows))
    _autosize(ws, min_w=12, max_w=80)
